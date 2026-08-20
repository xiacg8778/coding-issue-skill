#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量提单脚本：读取补充后的缺陷 Excel → 逐条在 CODING 创建 Bug 单
与 export_defects_excel.py 配套（列结构见该脚本 COLUMNS）。

用法:
  python batch_create_bugs.py 缺陷.xlsx                    # dry-run：全量校验，零写入
  python batch_create_bugs.py 缺陷.xlsx --execute          # 真实创建
  python batch_create_bugs.py 缺陷.xlsx --execute --write-back  # 创建后回填 issue_code/url/status 到 Excel

行为约定:
  - create 列为 no 的行跳过；必填（project/assignee）缺失的行标 skipped 并给出原因，不中断其他行
  - 单条创建失败不中断批次，结束时输出汇总表；--write-back 时失败行 status 写入失败原因
  - token 默认读 ~/.workbuddy/mcp.json（同 create_bug.py）
"""
import argparse
import datetime
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("错误: 缺少 openpyxl。安装: pip install openpyxl")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from create_bug import (  # noqa: E402  复用已验证的 API 封装
    call_api, find_member, get_defect_fields, pick_option, upload_attachment, get_token,
)

# 与 export_defects_excel.py 的 COLUMNS 对齐
HDR = ["defect_id", "title*", "description", "severity", "project*", "assignee*",
       "owner", "category", "images", "priority", "due_date", "create*",
       "issue_code", "issue_url", "status"]


def read_rows(xlsx_path):
    """读取「提单清单」sheet，返回 [{表头: 值}]，表头以第1行文字匹配（容忍列序变化）"""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["提单清单"]
    header = [str(c.value or "").strip() for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    missing = [h for h in ("project*", "assignee*", "title*", "create*") if h not in idx]
    if missing:
        sys.exit(f"错误: Excel 缺少必需列 {missing}，请用 export_defects_excel.py 生成的模板")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        row = {h: ("" if i >= len(r) or r[i] is None else str(r[i]).strip()) for h, i in idx.items()}
        row["_row_num"] = len(rows) + 2
        rows.append(row)
    if not rows:
        sys.exit("错误: Excel 无数据行")
    return rows, wb


def resolve_team_host(token, project):
    """经 DescribeProjectByName → DescribeTeam 动态解析团队域名。
    ⚠️ 响应结构：团队信息在 Data 键下（Data.TeamHost），不是 Team 键——读错键会误判'接口不可用'。
    解析失败返回 None，调用方回退 --base-url 或默认值。"""
    try:
        proj = call_api(token, {"Action": "DescribeProjectByName", "ProjectName": project})["Project"]
        team_id = proj.get("TeamId")
        if not team_id:
            return None
        t = call_api(token, {"Action": "DescribeTeam", "TeamId": team_id})
        data = t.get("Data") or {}
        return data.get("TeamHost")
    except SystemExit:
        return None


def issue_url(project, code, base=None):
    """生成事项页直达 URL。
    base 优先级：显式 --base-url > resolve_team_host 动态解析 > 默认 sugaoxin.coding.net（本团队实证值）。
    路径格式实证：https://sugaoxin.coding.net/p/<project>/bug-tracking/issues/<code>（用户验证可访问）。"""
    b = (base or "https://sugaoxin.coding.net").rstrip("/")
    return f"{b}/p/{project}/bug-tracking/issues/{code}"


def server_side_dedup(token, title, project):
    """按标题在项目 DEFECT 列表查同名单（最近 50 条），返回 (code, name) 或 None。
    用于"网络模糊失败"行的重提前防重：请求可能已送达建单，只是响应丢失。"""
    try:
        resp = call_api(token, {"Action": "DescribeIssueList", "ProjectName": project,
                                "IssueType": "DEFECT", "PageNumber": 1, "PageSize": 50})
        for i in resp.get("IssueList", []):
            if (i.get("Name") or "").strip() == (title or "").strip():
                return str(i.get("Code")), i.get("Name")
    except SystemExit:
        pass  # 查询失败不阻断主流程，退回正常创建
    return None


def validate_row(row, known_projects):
    """dry-run 校验：返回错误列表（空=可提单）"""
    errs = []
    if row.get("create*", "").lower() not in ("yes", "y", "true", "1"):
        return ["create≠yes 跳过"]
    for f in ("project*", "assignee*", "title*"):
        if not row.get(f):
            errs.append(f"缺必填 {f}")
    p = row.get("priority", "")
    if p and p not in ("0", "1", "2", "3"):
        errs.append(f"priority 非法({p})，应为0-3")
    dd = row.get("due_date", "")
    if dd and (len(dd) != 10 or dd[4] != "-" or dd[7] != "-"):
        errs.append(f"due_date 格式({dd})应为 YYYY-MM-DD")
    for img in [x for x in row.get("images", "").split(";") if x.strip()]:
        if not os.path.exists(img.strip()):
            errs.append(f"图片不存在 {img.strip()}")
    if known_projects is not None and row.get("project*") and row["project*"] not in known_projects:
        errs.append(f"项目在 token 可见范围外: {row['project*']}")
    return errs


def main():
    ap = argparse.ArgumentParser(description="CODING 批量提单（Excel 驱动）")
    ap.add_argument("xlsx", help="补充后的缺陷 Excel")
    ap.add_argument("--execute", action="store_true", help="真实创建（缺省仅 dry-run 校验）")
    ap.add_argument("--write-back", action="store_true", help="把结果回填 Excel")
    ap.add_argument("--token", default=None)
    ap.add_argument("--base-url", default=None, help="团队域名（如 https://xxx.coding.net），用于生成可直达的事项页 URL；团队域名 API 不可查，须从浏览器地址栏复制")
    args = ap.parse_args()

    rows, wb = read_rows(args.xlsx)
    todo = [r for r in rows if r.get("create*", "").lower() in ("yes", "y", "true", "1")]
    print(f"共 {len(rows)} 行，待提单 {len(todo)} 行，跳过 {len(rows) - len(todo)} 行 | 模式: {'EXECUTE' if args.execute else 'DRY-RUN'}")

    token = get_token(args.token)

    # 预取 token 可见项目清单（只读一次，用于校验；DescribeUserProjects 必须带 UserId）
    known_projects = None
    try:
        me = call_api(token, {"Action": "DescribeCodingCurrentUser"})
        uid = me.get("User", {}).get("Id")
        if uid:
            resp = call_api(token, {"Action": "DescribeUserProjects", "UserId": uid})
            known_projects = [p.get("Name") for p in resp.get("ProjectList", []) if p.get("Name")]
            print(f"[预检] token 可见项目 {len(known_projects)} 个")
        else:
            print("[预检] 无法获取当前用户 Id，跳过项目存在性校验")
    except SystemExit as e:
        print(f"[预检] 项目清单获取失败（{str(e)[:80]}），跳过项目存在性校验")

    # 缓存: 项目 → (成员表/字段配置)，避免同项目多行重复查询
    proj_cache = {}

    # 团队域名动态解析（一次）：DescribeProjectByName → DescribeTeam → Data.TeamHost
    team_host = None
    if todo:
        team_host = resolve_team_host(token, todo[0].get("project*") or "")
        if team_host:
            print(f"[域名] TeamHost 解析成功: {team_host}")
        else:
            print("[域名] TeamHost 解析失败，回退 --base-url/默认值")
    effective_base = args.base_url or team_host

    results = []
    for n, row in enumerate(todo, 1):
        did = row.get("defect_id") or f"行{row['_row_num']}"
        # 幂等保护①：Excel 已回填 issue_code 的行跳过，防重跑重复建单
        if row.get("issue_code"):
            results.append({"did": did, "status": "already_created", "detail": f"issue_code={row['issue_code']} 已存在，跳过"})
            print(f"  [{n}/{len(todo)}] {did} ⏭ 已创建过 {row['issue_code']}，跳过")
            continue
        # 幂等保护②：失败/跳过行重提时先按标题查服务端，防"网络模糊失败"重复建单
        # 背景：CreateIssue 请求送达但响应被 SSL 瞬断吞掉时，客户端误判 failed，重提即重复
        if str(row.get("status", "")).startswith(("failed", "skipped")):
            exist = server_side_dedup(token, row["title*"], row.get("project*") or "")
            if exist:
                results.append({"did": did, "status": "recovered", "code": exist[0],
                                "url": issue_url(row.get("project*") or "", exist[0], effective_base),
                                "detail": f"服务端已存在同名单 {exist[0]}（此前模糊失败实际已建单），直接认领"})
                print(f"  [{n}/{len(todo)}] {did} ↻ 服务端已有 {exist[0]}，认领（不重建）")
                continue
        errs = validate_row(row, known_projects)
        if errs:
            results.append({"did": did, "status": "skipped", "detail": "; ".join(errs)})
            print(f"  [{n}/{len(todo)}] {did} ✗ skipped: {'; '.join(errs)}")
            continue
        if not args.execute:
            results.append({"did": did, "status": "dry_run_ok", "detail": "校验通过，--execute 后创建"})
            print(f"  [{n}/{len(todo)}] {did} ✓ dry-run 通过（{row['project*']} / {row['assignee*']} / {row['title*'][:30]}）")
            continue

        project = row["project*"]
        try:
            if project not in proj_cache:
                resp = call_api(token, {"Action": "DescribeProjectByName", "ProjectName": project})
                pid = resp.get("Project", {}).get("Id")
                if not pid:
                    raise RuntimeError(f"项目不存在: {project}")
                fields = get_defect_fields(token, project)
                proj_cache[project] = (pid, fields, {})
            pid, fields, member_cache = proj_cache[project]

            # 成员解析（带缓存）
            def member_id(name):
                if name not in member_cache:
                    member_cache[name] = find_member(token, pid, name)
                return member_cache[name]

            assignee_id = member_id(row["assignee*"])
            owner_name = row.get("owner") or row["assignee*"]
            owner_id = member_id(owner_name)

            # 必填自定义字段补齐（同 create_bug.py 逻辑）
            custom_values = [{"Id": fields["问题归属人"]["field_id"], "Content": str(owner_id)}]
            defect_type_id = None
            if "缺陷类型" in fields:
                defect_type_id = int(pick_option(fields["缺陷类型"], prefer="功能缺陷"))
            category_val = None
            if "Bug归类" in fields:
                category_val = pick_option(fields["Bug归类"], prefer=row.get("category") or None)
                if row.get("category") and category_val is None:
                    raise RuntimeError(f"Bug归类无「{row['category']}」选项")
            if category_val:
                custom_values.append({"Id": fields["Bug归类"]["field_id"], "Content": str(category_val)})
            for fname, f in fields.items():
                if fname in ("处理人", "缺陷类型", "问题归属人", "Bug归类", "关注人", "优先级", "截止日期", "开始日期", "开发截止日期", "进度"):
                    continue
                if f["required"] and f["component_type"] in ("SELECT_SINGLE", "SELECT_MULTI"):
                    val = pick_option(f, prefer={"严重程度": "一般", "问题来源": "测试环境"}.get(fname))
                    if val:
                        custom_values.append({"Id": f["field_id"], "Content": str(val)})

            due = row.get("due_date") or (datetime.date.today() + datetime.timedelta(days=7)).isoformat()
            resp = call_api(token, {
                "Action": "CreateIssue", "ProjectName": project, "Type": "DEFECT",
                "Name": row["title*"], "Priority": row.get("priority") or "1",
                "AssigneeId": assignee_id, "Description": row.get("description") or row["title*"],
                "DueDate": due,
                "CustomFieldValues": custom_values, **({"DefectTypeId": defect_type_id} if defect_type_id else {}),
            })
            issue_code = resp["Issue"]["Code"]

            # 附件（可选）
            file_ids, attached = [], 0
            for img in [x.strip() for x in row.get("images", "").split(";") if x.strip()]:
                file_ids.append(upload_attachment(token, project, img))
                attached += 1
            if file_ids:
                call_api(token, {"Action": "ModifyIssue", "ProjectName": project,
                                 "IssueCode": issue_code, "Name": row["title*"], "FileIds": file_ids})

            # 创建后验证（同 create_bug.py 契约）
            final = call_api(token, {"Action": "DescribeIssue", "ProjectName": project, "IssueCode": issue_code})["Issue"]
            assert final["Assignee"]["Name"] and len(final.get("Files", [])) >= attached

            url = issue_url(project, str(issue_code), effective_base)
            results.append({"did": did, "status": "created", "code": str(issue_code), "url": url, "detail": f"附件{attached}"})
            print(f"  [{n}/{len(todo)}] {did} ✓ created: {issue_code}（附件{attached}）")

        except SystemExit as e:
            results.append({"did": did, "status": "failed", "detail": str(e)[:200]})
            print(f"  [{n}/{len(todo)}] {did} ✗ failed: {str(e)[:200]}")
        except Exception as e:
            results.append({"did": did, "status": "failed", "detail": f"{type(e).__name__}: {e}"[:200]})
            print(f"  [{n}/{len(todo)}] {did} ✗ failed: {type(e).__name__}: {e}"[:200])

    # 汇总
    created = sum(1 for r in results if r["status"] == "created")
    ok = sum(1 for r in results if r["status"] == "dry_run_ok")
    failed = sum(1 for r in results if r["status"] == "failed")
    skipped = sum(1 for r in results if r["status"] == "skipped")
    print(f"\n{'='*60}\n汇总: created={created} dry_run_ok={ok} failed={failed} skipped={skipped}")
    if failed:
        print("失败明细:")
        for r in results:
            if r["status"] == "failed":
                print(f"  {r['did']}: {r['detail']}")

    # 回填
    if args.write_back:
        ws = wb["提单清单"]
        header = [str(c.value or "").strip() for c in ws[1]]
        col = {h: i + 1 for i, h in enumerate(header)}
        by_did = {r["did"]: r for r in results}
        written = 0
        for row in ws.iter_rows(min_row=2):
            did = str(row[col["defect_id"] - 1].value or "").strip()
            if not did:
                continue
            r = by_did.get(did)
            if not r:
                continue
            row[col["issue_code"] - 1].value = r.get("code", "")
            row[col["issue_url"] - 1].value = r.get("url", "")
            row[col["status"] - 1].value = r["status"] + (f" ({r['detail']})" if r["status"] in ("failed", "skipped") else "")
            written += 1
        wb.save(args.xlsx)
        print(f"回填完成: {written} 行 → {args.xlsx}")


if __name__ == "__main__":
    main()
