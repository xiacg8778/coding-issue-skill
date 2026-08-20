#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
缺陷台账 → Excel 补充模板导出
读取 qa-team 的 defect_ledger.json（或 qa-reporter 缺陷草稿列表），生成供用户补充提单信息的 xlsx。

用法:
  python export_defects_excel.py <defect_ledger.json> [-o 输出.xlsx] [--project 默认项目名]
  不指定 -o 时默认输出到输入文件同目录 defects_<run_id>.xlsx
  --project 预填 project 列（默认 BrainServicePlatform）

Excel 结构:
  - Sheet「提单清单」: QA 产出列（只读灰底）+ 用户补充列（黄底，标 * 必填）
  - Sheet「说明」: 填写指引、列说明、下游 batch_create_bugs.py 用法
"""
import argparse
import datetime
import hashlib
import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("错误: 缺少 openpyxl。安装: pip install openpyxl")

# 列定义: (表头, 宽度, 来源, 必填)
# 来源=ledger 表示从缺陷台账自动带出（灰底只读）；=user 表示用户补充（黄底）
COLUMNS = [
    ("defect_id",       14, "ledger", False),
    ("title*",          40, "ledger", True),   # 台账带出，可改
    ("description",     50, "ledger", False),  # 台账带出，可改
    ("severity",        10, "ledger", False),  # 台带分类，供人参考
    ("project*",        24, "user",   True),
    ("assignee*",       12, "user",   True),
    ("owner",           12, "user",   False),  # 空=同处理人
    ("category",        12, "user",   False),  # Bug归类选项标题，空=第一个选项
    ("images",          36, "user",   False),  # 多图分号分隔
    ("priority",        8,  "user",   False),  # 0低 1中 2高 3紧急，默认1
    ("due_date",        12, "user",   False),  # YYYY-MM-DD，默认7天后
    ("create*",         8,  "user",   True),   # yes/no 是否提单
    ("issue_code",      12, "result", False),  # 提单结果回填
    ("issue_url",       34, "result", False),
    ("status",          20, "result", False),
]

GRAY = PatternFill("solid", fgColor="ECECEC")
YELLOW = PatternFill("solid", fgColor="FFF6D6")
GREEN = PatternFill("solid", fgColor="E2EFDA")


def load_defects(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    # 兼容两种形态: qa-team 台账 {run_id, defects:[...]} 或裸列表 [...]
    defects = data.get("defects") if isinstance(data, dict) else data
    if not defects:
        sys.exit("错误: 输入文件中没有缺陷条目（defects 为空）")
    return data.get("run_id", "") if isinstance(data, dict) else "", defects


def collect_screenshots(run_root, case_id, seen_hashes):
    """收集用例证据截图（evidence/*/case_id/ 下递归 png/jpg），按 md5 跨缺陷去重。
    返回 (绝对路径列表, 本用例新增数)；重复图（同 md5 已被其他缺陷引用）跳过。"""
    cands = []
    for side in ("web", "api", "mobile"):
        cdir = os.path.join(run_root, "evidence", side, case_id)
        if os.path.isdir(cdir):
            for dp, _, fs in os.walk(cdir):
                for f in sorted(fs):
                    if f.lower().endswith((".png", ".jpg", ".jpeg")):
                        cands.append(os.path.join(dp, f))
    kept, dup = [], 0
    for p in sorted(cands):
        h = hashlib.md5(open(p, "rb").read()).hexdigest()
        if h in seen_hashes:
            dup += 1
            continue
        seen_hashes.add(h)
        kept.append(p)
    if dup:
        print(f"    [去重] {case_id}: 跳过 {dup} 张跨缺陷重复截图")
    return kept


def load_case_steps(run_root, case_id):
    """从 test_cases.json 提取用例复现步骤与预期，返回 (steps, expected) 或 (None, None)"""
    tc = os.path.join(run_root, "test_cases.json")
    if not os.path.isfile(tc):
        return None, None
    try:
        with open(tc, encoding="utf-8") as f:
            data = json.load(f)
        for c in (data.get("cases") or [] if isinstance(data, dict) else data):
            if c.get("case_id") == case_id:
                return c.get("steps"), c.get("expected")
    except Exception:
        pass
    return None, None


def main():
    ap = argparse.ArgumentParser(description="缺陷台账导出 Excel 补充模板")
    ap.add_argument("ledger", help="defect_ledger.json 路径（qa-team Phase 6 产出）")
    ap.add_argument("-o", "--output", default=None, help="输出 xlsx 路径（默认同目录 defects_<run_id>.xlsx）")
    ap.add_argument("--project", default="BrainServicePlatform", help="project 列预填默认值（默认 BrainServicePlatform，留空字符串则不预填）")
    ap.add_argument("--run-root", default=None, help="qa-team run 目录（自动关联 test_cases.json 复现步骤与 evidence 截图，md5 去重）")
    ap.add_argument("--max-images", type=int, default=4, help="每条缺陷最多挂载截图数（默认 4）")
    args = ap.parse_args()

    run_id, defects = load_defects(args.ledger)
    # run 目录自动推断：缺省取 ledger 同级向上两级（report/defect_ledger.json 的惯例位置）
    run_root = args.run_root
    if not run_root:
        guess = os.path.dirname(os.path.dirname(os.path.abspath(args.ledger)))
        if os.path.isfile(os.path.join(guess, "test_cases.json")):
            run_root = guess
            print(f"[自动关联] run 目录: {run_root}")
    seen_hashes = set()
    out = args.output or os.path.join(
        os.path.dirname(os.path.abspath(args.ledger)),
        f"defects_{run_id or 'draft'}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "提单清单"

    header_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(vertical="top", wrap_text=True)

    # 表头（第1行）+ 颜色图例
    for col, (title, width, src, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.alignment = center
        c.fill = {"ledger": GRAY, "user": YELLOW, "result": GREEN}[src]
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    # 数据行
    for i, d in enumerate(defects, 2):
        title = d.get("title") or f"[{d.get('defect_id','')}] {d.get('case_id','')} 用例失败"
        # 描述 = 摘要 + 复现步骤 + 预期 + 回归归因（若有）
        desc_parts = []
        for k in ("summary", "description", "evidence"):
            if d.get(k):
                desc_parts.append(str(d[k]))
        cid = d.get("case_id")
        steps, expected = (None, None)
        if run_root and cid:
            steps, expected = load_case_steps(run_root, cid)
        if steps:
            desc_parts.append(f"【复现步骤】\n{steps}")
        if expected:
            desc_parts.append(f"【预期结果】\n{expected}")
        rg = d.get("regression")
        if isinstance(rg, dict):
            if rg.get("root_cause"):
                desc_parts.append(f"【根因】{rg['root_cause']}")
            if rg.get("evidence"):
                desc_parts.append(f"【证据】{rg['evidence']}")
            if rg.get("reproduced"):
                desc_parts.append("（已复测复现并经后端确认）")
        # 证据截图自动关联（跨缺陷 md5 去重）
        images = []
        if run_root and cid:
            images = collect_screenshots(run_root, cid, seen_hashes)[:args.max_images]
        row_vals = {
            "defect_id": d.get("defect_id", ""),
            "title*": title,
            "description": "\n\n".join(desc_parts),
            "severity": d.get("severity", ""),
            "project*": args.project,
            "images": ";".join(images),
            "create*": "yes",
        }
        if images:
            print(f"    {d.get('defect_id','')}: 关联 {len(images)} 张证据截图")
        for col, (title_h, _, src, _) in enumerate(COLUMNS, 1):
            c = ws.cell(row=i, column=col, value=row_vals.get(title_h, ""))
            c.alignment = center if title_h in ("defect_id", "severity", "priority", "due_date", "create*", "issue_code") else left_wrap
            c.fill = {"ledger": GRAY, "user": YELLOW, "result": GREEN}[src]
        # 默认值提示: owner 空则同 assignee 由下游脚本处理，这里不预填避免误导

    # 说明 Sheet
    ws2 = wb.create_sheet("说明")
    notes = [
        ["缺陷提单补充模板 — 填写说明"],
        [f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}    来源: {os.path.basename(args.ledger)}    run_id: {run_id or '-'}"],
        [""],
        ["颜色: 灰底=QA产出(可改) | 黄底=用户补充(标*必填) | 绿底=提单结果(勿动，脚本回填)"],
        [""],
        ["用户补充列说明:"],
        ["  project*   CODING 项目名称（已预填，可改）"],
        ["  assignee*  处理人姓名（须为项目成员）"],
        ["  owner      问题归属人，留空=同处理人"],
        ["  category   Bug归类选项标题（如 web前端），留空=取第一个选项"],
        ["  images     本地图片绝对路径，多张用分号 ; 分隔"],
        ["  priority   0低 1中 2高 3紧急，留空默认1(中)"],
        ["  due_date   截止日期 YYYY-MM-DD，留空默认7天后"],
        ["  create*    yes=提单，no=跳过"],
        [""],
        ["提单执行（先预检后真实）:"],
        ["  python ~/.workbuddy/skills/coding-issue-bug/scripts/batch_create_bugs.py <本文件.xlsx>            # dry-run 校验"],
        ["  python ~/.workbuddy/skills/coding-issue-bug/scripts/batch_create_bugs.py <本文件.xlsx> --execute  # 真实创建"],
        ["  追加 --write-back 可把 issue_code/url/status 回填到本 Excel"],
        [""],
        ["token 自动读取: ~/.workbuddy/mcp.json → mcpServers.coding-devops.env.CODING_TOKEN"],
    ]
    for r, row in enumerate(notes, 1):
        ws2.cell(row=r, column=1, value=row[0])
        if r == 1:
            ws2.cell(row=r, column=1).font = Font(bold=True, size=12)
    ws2.column_dimensions["A"].width = 110

    wb.save(out)
    print(f"✅ 已导出 {len(defects)} 条缺陷 → {out}")
    print(f"   用户补充黄底列后，先 dry-run 校验再 --execute 提单")


if __name__ == "__main__":
    main()
